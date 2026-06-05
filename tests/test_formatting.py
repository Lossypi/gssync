import pytest

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from gssync.formatting import (
    CellFormat,
    SheetFormatting,
    color_gs_to_hex,
    color_hex_to_gs,
    color_hex_to_argb,
    color_argb_to_hex,
    px_to_char_width,
    char_width_to_px,
    px_to_points,
    points_to_px,
    gs_number_type,
    GS_TO_OPENPYXL_BORDER,
    OPENPYXL_TO_GS_BORDER,
    gs_cell_to_ir,
    ir_to_gs_cell_format,
    gs_grid_to_sheet_formatting,
    ir_to_gs_requests,
    openpyxl_cell_to_ir,
    apply_ir_to_openpyxl_cell,
)


# ── dataclasses + helpers ──────────────────────────────────────────────────────

def test_cellformat_is_empty_true_for_all_none():
    assert CellFormat().is_empty() is True


def test_cellformat_is_empty_false_when_any_set():
    assert CellFormat(bold=True).is_empty() is False


def test_sheetformatting_defaults_are_independent():
    a = SheetFormatting()
    b = SheetFormatting()
    a.cells[(0, 0)] = CellFormat(bold=True)
    assert b.cells == {}


def test_color_gs_to_hex_full_channels():
    assert color_gs_to_hex({"red": 1.0, "green": 0.0, "blue": 0.0}) == "#FF0000"


def test_color_gs_to_hex_missing_channel_defaults_zero():
    assert color_gs_to_hex({"red": 1.0}) == "#FF0000"


def test_color_hex_to_gs_roundtrip():
    gs = color_hex_to_gs("#FF8000")
    assert gs["red"] == pytest.approx(1.0)
    assert gs["green"] == pytest.approx(128 / 255)
    assert gs["blue"] == pytest.approx(0.0)


def test_color_hex_to_argb():
    assert color_hex_to_argb("#1A2B3C") == "FF1A2B3C"


def test_color_argb_to_hex():
    assert color_argb_to_hex("FF1A2B3C") == "#1A2B3C"


def test_width_conversion_roundtrip_is_close():
    px = char_width_to_px(8.43)
    assert abs(px_to_char_width(px) - 8.43) < 0.5


def test_char_width_to_px_known_value():
    assert char_width_to_px(8.43) == round(8.43 * 7 + 5)


def test_height_conversion_roundtrip():
    assert points_to_px(px_to_points(20)) == 20


def test_gs_number_type_date():
    assert gs_number_type("dd.mm.yyyy") == "DATE"


def test_gs_number_type_percent():
    assert gs_number_type("0.0%") == "PERCENT"


def test_gs_number_type_plain_number():
    assert gs_number_type("0.00") == "NUMBER"


def test_border_maps_are_inverse_for_common_styles():
    assert GS_TO_OPENPYXL_BORDER["SOLID"] == "thin"
    assert OPENPYXL_TO_GS_BORDER["thin"] == "SOLID"


# ── Google Sheets JSON ↔ IR ────────────────────────────────────────────────────

def test_gs_cell_to_ir_reads_text_format():
    value = {"userEnteredFormat": {"textFormat": {"bold": True, "fontSize": 14,
             "foregroundColor": {"red": 1.0}}}}
    cf = gs_cell_to_ir(value)
    assert cf.bold is True
    assert cf.font_size == 14
    assert cf.text_color == "#FF0000"


def test_gs_cell_to_ir_reads_align_wrap_fill():
    value = {"userEnteredFormat": {
        "backgroundColor": {"blue": 1.0},
        "horizontalAlignment": "CENTER",
        "verticalAlignment": "MIDDLE",
        "wrapStrategy": "WRAP",
    }}
    cf = gs_cell_to_ir(value)
    assert cf.background_color == "#0000FF"
    assert cf.h_align == "center"
    assert cf.v_align == "middle"
    assert cf.wrap is True


def test_gs_cell_to_ir_empty_value():
    assert gs_cell_to_ir({}).is_empty() is True


def test_ir_to_gs_cell_format_builds_fields():
    cf = CellFormat(bold=True, background_color="#0000FF")
    uef, fields = ir_to_gs_cell_format(cf)
    assert uef["textFormat"]["bold"] is True
    assert uef["backgroundColor"] == color_hex_to_gs("#0000FF")
    assert "textFormat.bold" in fields
    assert "backgroundColor" in fields


def test_ir_to_gs_cell_format_empty_has_no_fields():
    _, fields = ir_to_gs_cell_format(CellFormat())
    assert fields == []


def test_gs_grid_to_sheet_formatting_cells_and_dims():
    sheet = {"data": [{
        "rowData": [
            {"values": [{"userEnteredFormat": {"textFormat": {"bold": True}}}, {}]},
            {"values": [{}, {"userEnteredFormat": {"textFormat": {"italic": True}}}]},
        ],
        "columnMetadata": [{"pixelSize": 100}, {"pixelSize": 150}],
        "rowMetadata": [{"pixelSize": 21}, {"pixelSize": 30}],
    }]}
    sf = gs_grid_to_sheet_formatting(sheet)
    assert sf.cells[(0, 0)].bold is True
    assert sf.cells[(1, 1)].italic is True
    assert (0, 1) not in sf.cells
    assert sf.column_widths == {0: 100, 1: 150}
    assert sf.row_heights == {0: 21, 1: 30}


def test_ir_to_gs_requests_builds_repeatcell_and_dims():
    sf = SheetFormatting(
        cells={(2, 3): CellFormat(bold=True)},
        column_widths={1: 120},
        row_heights={4: 30},
    )
    requests = ir_to_gs_requests(99, sf)
    repeat = [r for r in requests if "repeatCell" in r][0]
    rng = repeat["repeatCell"]["range"]
    assert rng == {"sheetId": 99, "startRowIndex": 2, "endRowIndex": 3,
                   "startColumnIndex": 3, "endColumnIndex": 4}
    assert repeat["repeatCell"]["cell"]["userEnteredFormat"]["textFormat"]["bold"] is True
    dims = [r for r in requests if "updateDimensionProperties" in r]
    assert any(d["updateDimensionProperties"]["range"]["dimension"] == "COLUMNS" for d in dims)
    assert any(d["updateDimensionProperties"]["range"]["dimension"] == "ROWS" for d in dims)


def test_ir_to_gs_requests_skips_empty_cells():
    sf = SheetFormatting(cells={(0, 0): CellFormat()})
    assert ir_to_gs_requests(1, sf) == []


# ── openpyxl ↔ IR ──────────────────────────────────────────────────────────────

def test_openpyxl_cell_to_ir_reads_font_and_color():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.font = Font(name="Arial", size=14, bold=True, color="FFFF0000")
    cf = openpyxl_cell_to_ir(cell)
    assert cf.font_family == "Arial"
    assert cf.font_size == 14
    assert cf.bold is True
    assert cf.text_color == "#FF0000"


def test_openpyxl_cell_to_ir_reads_fill_align_wrap():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    cell.fill = PatternFill(fill_type="solid", fgColor="FF00FF00")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = openpyxl_cell_to_ir(cell)
    assert cf.background_color == "#00FF00"
    assert cf.h_align == "center"
    assert cf.v_align == "middle"
    assert cf.wrap is True


def test_openpyxl_cell_to_ir_skips_general_number_format():
    wb = openpyxl.Workbook()
    cf = openpyxl_cell_to_ir(wb.active["A1"])
    assert cf.number_format is None


def test_apply_ir_to_openpyxl_cell_roundtrip():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["A1"]
    src = CellFormat(font_family="Calibri", font_size=12, bold=True, italic=True,
                     text_color="#112233", background_color="#445566",
                     h_align="right", v_align="top", wrap=True, number_format="0.00")
    apply_ir_to_openpyxl_cell(cell, src)
    back = openpyxl_cell_to_ir(cell)
    assert back.font_family == "Calibri"
    assert back.bold is True
    assert back.italic is True
    assert back.text_color == "#112233"
    assert back.background_color == "#445566"
    assert back.h_align == "right"
    assert back.v_align == "top"
    assert back.wrap is True
    assert back.number_format == "0.00"


def test_apply_ir_to_openpyxl_cell_empty_is_noop():
    wb = openpyxl.Workbook()
    cell = wb.active["A1"]
    apply_ir_to_openpyxl_cell(cell, CellFormat())
    assert cell.font.name in (None, "Calibri")  # default font untouched


def test_apply_ir_border_roundtrip():
    wb = openpyxl.Workbook()
    cell = wb.active["A1"]
    apply_ir_to_openpyxl_cell(cell, CellFormat(borders={"top": {"style": "thin", "color": "#000000"}}))
    back = openpyxl_cell_to_ir(cell)
    assert back.borders["top"]["style"] == "thin"
