import json
from pathlib import Path
import pytest
from gssync.storage import read_local, write_local, list_local_sheets

SAMPLE = {
    "Sheet1": [["Name", "Age"], ["Alice", "30"], ["Bob", "25"]],
    "Sheet2": [["X", "Y"], ["1", "2"]],
}

def test_xlsx_roundtrip(tmp_path):
    path = tmp_path / "test.xlsx"
    write_local(path, "xlsx", SAMPLE)
    result = read_local(path, "xlsx")
    assert result["Sheet1"] == SAMPLE["Sheet1"]
    assert result["Sheet2"] == SAMPLE["Sheet2"]

def test_xlsx_preserves_existing_sheets(tmp_path):
    path = tmp_path / "test.xlsx"
    write_local(path, "xlsx", SAMPLE)
    write_local(path, "xlsx", {"Sheet3": [["A"], ["1"]]})
    result = read_local(path, "xlsx")
    assert "Sheet1" in result
    assert "Sheet3" in result

def test_json_roundtrip(tmp_path):
    path = tmp_path / "test.json"
    write_local(path, "json", SAMPLE)
    result = read_local(path, "json")
    assert result == SAMPLE

def test_csv_roundtrip(tmp_path):
    path = tmp_path / "csv_dir"
    write_local(path, "csv", SAMPLE)
    result = read_local(path, "csv")
    assert result["Sheet1"] == SAMPLE["Sheet1"]
    assert result["Sheet2"] == SAMPLE["Sheet2"]

def test_list_local_sheets_xlsx(tmp_path):
    path = tmp_path / "test.xlsx"
    write_local(path, "xlsx", SAMPLE)
    names = list_local_sheets(path, "xlsx")
    assert set(names) == {"Sheet1", "Sheet2"}

def test_list_local_sheets_missing_file(tmp_path):
    path = tmp_path / "missing.xlsx"
    assert list_local_sheets(path, "xlsx") == []


# ── formatting ──────────────────────────────────────────────────────────────

import openpyxl
from openpyxl.styles import Font, PatternFill

from gssync.formatting import SheetFormatting, CellFormat
from gssync.storage import read_xlsx_formatting, apply_formatting_to_xlsx


def _make_fmt_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    ws["A1"].font = Font(bold=True, color="FFFF0000")
    ws.column_dimensions["A"].width = 20
    ws.row_dimensions[1].height = 30
    wb.save(path)


def test_read_xlsx_formatting_reads_cells_and_dims(tmp_path):
    xlsx = tmp_path / "data.xlsx"
    _make_fmt_xlsx(xlsx)
    sf = read_xlsx_formatting(xlsx, "Sheet1")
    assert sf.cells[(0, 0)].bold is True
    assert sf.cells[(0, 0)].text_color == "#FF0000"
    assert 0 in sf.column_widths
    assert 0 in sf.row_heights


def test_read_xlsx_formatting_missing_sheet_returns_empty(tmp_path):
    xlsx = tmp_path / "data.xlsx"
    _make_fmt_xlsx(xlsx)
    sf = read_xlsx_formatting(xlsx, "Missing")
    assert sf.cells == {}


def test_apply_formatting_to_xlsx_writes_styles(tmp_path):
    xlsx = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.active["A1"] = "x"
    wb.save(xlsx)
    sf = SheetFormatting(
        cells={(0, 0): CellFormat(bold=True, background_color="#00FF00")},
        column_widths={0: 150},
        row_heights={0: 40},
    )
    apply_formatting_to_xlsx(xlsx, "Sheet1", sf)
    back = read_xlsx_formatting(xlsx, "Sheet1")
    assert back.cells[(0, 0)].bold is True
    assert back.cells[(0, 0)].background_color == "#00FF00"
    assert 0 in back.column_widths
