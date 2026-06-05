import csv
import json
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

SheetData = Dict[str, List[List]]


def _normalize_xlsx_cell(value: object) -> object:
    if value is None or isinstance(value, (str, int, float)):
        return value
    if hasattr(value, "text"):
        return value.text
    return str(value)


def read_xlsx(path: Path) -> SheetData:
    wb = openpyxl.load_workbook(path)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        result[name] = [[_normalize_xlsx_cell(cell.value) for cell in row] for row in ws.iter_rows()]
    return result


def write_xlsx(path: Path, data: SheetData) -> None:
    wb = openpyxl.load_workbook(path) if path.exists() else openpyxl.Workbook()
    if not path.exists() and wb.sheetnames == ["Sheet"]:
        del wb["Sheet"]
    for name, rows in data.items():
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append([v if v is not None else "" for v in row])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def read_json(path: Path) -> SheetData:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data: SheetData) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def read_csv_dir(path: Path) -> SheetData:
    result = {}
    for csv_file in sorted(path.glob("*.csv")):
        with open(csv_file, newline="", encoding="utf-8") as f:
            result[csv_file.stem] = list(csv.reader(f))
    return result


def write_csv_dir(path: Path, data: SheetData) -> None:
    path.mkdir(parents=True, exist_ok=True)
    for name, rows in data.items():
        with open(path / f"{name}.csv", "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)


def read_local(path: Path, fmt: str) -> SheetData:
    if fmt == "xlsx":
        return read_xlsx(path)
    if fmt == "json":
        return read_json(path)
    if fmt == "csv":
        return read_csv_dir(path)
    raise ValueError(f"Unknown format: {fmt}")


def write_local(path: Path, fmt: str, data: SheetData) -> None:
    if fmt == "xlsx":
        write_xlsx(path, data)
    elif fmt == "json":
        write_json(path, data)
    elif fmt == "csv":
        write_csv_dir(path, data)
    else:
        raise ValueError(f"Unknown format: {fmt}")


def list_local_sheets(path: Path, fmt: str) -> List[str]:
    exists = (path.exists() and path.is_dir()) if fmt == "csv" else path.exists()
    if not exists:
        return []
    return list(read_local(path, fmt).keys())


def read_xlsx_formatting(path: Path, sheet_name: str):
    from .formatting import (
        openpyxl_cell_to_ir, char_width_to_px, points_to_px, SheetFormatting,
    )
    wb = openpyxl.load_workbook(path)
    if sheet_name not in wb.sheetnames:
        return SheetFormatting()
    ws = wb[sheet_name]
    sf = SheetFormatting()
    for row in ws.iter_rows():
        for cell in row:
            # Skip cells that were never explicitly styled. openpyxl hands every
            # cell a default Font(Calibri, 11), so without this guard every blank
            # cell would be captured and pushed to GS.
            if not cell.has_style:
                continue
            cf = openpyxl_cell_to_ir(cell)
            if not cf.is_empty():
                sf.cells[(cell.row - 1, cell.column - 1)] = cf
    for letter, dim in ws.column_dimensions.items():
        if dim.width is not None:
            sf.column_widths[column_index_from_string(letter) - 1] = char_width_to_px(dim.width)
    for idx, dim in ws.row_dimensions.items():
        if dim.height is not None:
            sf.row_heights[idx - 1] = points_to_px(dim.height)
    return sf


def apply_formatting_to_xlsx(path: Path, sheet_name: str, sf) -> None:
    from .formatting import apply_ir_to_openpyxl_cell, px_to_char_width, px_to_points
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    for (r, c), cf in sf.cells.items():
        apply_ir_to_openpyxl_cell(ws.cell(row=r + 1, column=c + 1), cf)
    for col, px in sf.column_widths.items():
        ws.column_dimensions[get_column_letter(col + 1)].width = px_to_char_width(px)
    for row, px in sf.row_heights.items():
        ws.row_dimensions[row + 1].height = px_to_points(px)
    wb.save(path)
